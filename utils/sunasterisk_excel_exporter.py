"""
Sun Asterisk Excel Estimation Exporter
======================================

Generates Excel estimation sheets following Sun Asterisk standard format
with bilingual JP/EN headers, role-specific effort breakdown, and formulas.

Author: AI Assistant
Date: 2025-10-04
"""

import os
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Import logging
from utils.logger import get_logger

# Initialize logger
logger = get_logger(__name__)


class SunAsteriskExcelExporter:
    """
    Excel exporter following Sun Asterisk estimation sheet standard.

    Schema: 20 columns with bilingual headers, role × task type breakdown,
    MD per MM configuration, and formula-based calculations.
    """

    # Column definitions (order matters!)
    COLUMNS = [
        "No",                           # 1
        "Category",                     # 2
        "Parent Task",                  # 3
        "Sub Task",                     # 4
        "Sub.No",                       # 5
        "Premise",                      # 6
        "備考 Remark",                   # 7
        "Backend - Implement",          # 8
        "Backend - FixBug",             # 9
        "Backend - Unit Test",          # 10
        "Frontend - Implement",         # 11
        "Frontend - FixBug",            # 12
        "Frontend - Unit Test",         # 13
        "Frontend - Responsive",        # 14
        "Testing - Implement",          # 15
        "Total (MD)",                   # 16
        "Note"                          # 17
    ]

    # Effort column indices (1-based for Excel)
    EFFORT_COLUMNS = [8, 9, 10, 11, 12, 13, 14, 15]  # Backend-Implement through Testing-Implement
    TOTAL_COLUMN = 16  # Total (MD)

    # Minimum column widths (in Excel character units, ~7px per unit)
    # 50 character units ≈ 350px (half of previous 100)
    COLUMN_MIN_WIDTHS = {
        "Parent Task": 50,     # Column 3 - half width
        "Sub Task": 50,        # Column 4 - half width
        "Premise": 50,         # Column 6 - half width
        "備考 Remark": 50      # Column 7 - half width
    }

    def __init__(
        self,
        no: str = "001",
        version: str = "1.0",
        issue_date: Optional[str] = None,
        md_per_mm: int = 20
    ):
        """
        Initialize Sun Asterisk Excel Exporter.

        Args:
            no: Document number (e.g., "001")
            version: Document version (e.g., "1.0")
            issue_date: Issue date (defaults to today)
            md_per_mm: Man-days per man-month (default: 20)
        """
        self.no = no
        self.version = version
        self.issue_date = issue_date or datetime.now().strftime("%Y-%m-%d")
        self.md_per_mm = md_per_mm

    def export(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        project_id: str = None
    ) -> str:
        """
        Export estimation data to Sun Asterisk Excel format.

        Args:
            data: List of task dictionaries with estimation data
            filename: Output filename (auto-generated if None)
            project_id: Project identifier for project-scoped storage

        Returns:
            str: Path to exported Excel file
        """
        from config import Config
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"sunasterisk_estimation_{timestamp}.xlsx"

        # Determine output directory
        if project_id:
            output_dir = Config.get_project_result_dir(project_id)
            os.makedirs(output_dir, exist_ok=True)
            filepath = os.path.join(output_dir, filename)
        else:
            filepath = os.path.join(os.getcwd(), filename)

        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Estimation"

            # Build header
            self._build_header(ws)

            # Build data table (returns last row number)
            last_row = self._build_data_table(ws, data)

            # Add Total (MM) row
            self._add_total_mm_row(ws, last_row)

            # Apply formatting
            self._apply_formatting(ws, last_row)

            # Freeze panes (header + subheader)
            ws.freeze_panes = "A8"  # Freeze everything above row 8

            # Save workbook
            wb.save(filepath)

            logger.info(f"✅ Sun Asterisk Excel export completed: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"❌ Error exporting Sun Asterisk Excel: {e}")
            raise

    def _build_header(self, ws):
        """Build header section with company info and metadata."""
        # A1:F1 - "ESTIMATION"
        ws.merge_cells("A1:F1")
        ws["A1"] = "ESTIMATION"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # H1:J1 - "No: {no}"
        ws.merge_cells("H1:J1")
        ws["H1"] = f"No: {self.no}"
        ws["H1"].alignment = Alignment(horizontal="left", vertical="center")

        # H2:J2 - "Version: {version}"
        ws.merge_cells("H2:J2")
        ws["H2"] = f"Version: {self.version}"
        ws["H2"].alignment = Alignment(horizontal="left", vertical="center")

        # H3:J3 - "Issue Date: {issue_date}"
        ws.merge_cells("H3:J3")
        ws["H3"] = f"Issue Date: {self.issue_date}"
        ws["H3"].alignment = Alignment(horizontal="left", vertical="center")

        # A2:F2 - "SUN ASTERISK VIETNAM CO., LTD"
        ws.merge_cells("A2:F2")
        ws["A2"] = "SUN ASTERISK VIETNAM CO., LTD"
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # A3:F3 - "ISO/IEC 27001:2013 & ISO 9001:2015"
        ws.merge_cells("A3:F3")
        ws["A3"] = "ISO/IEC 27001:2013 & ISO 9001:2015"
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

        # L1 - "MD per MM"
        ws["L1"] = "MD per MM"
        ws["L1"].font = Font(bold=True)
        ws["L1"].alignment = Alignment(horizontal="right", vertical="center")

        # M1 - {md_per_mm}
        ws["M1"] = self.md_per_mm
        ws["M1"].font = Font(bold=True)
        ws["M1"].alignment = Alignment(horizontal="center", vertical="center")

        # Row 5: Role group headers
        # Backend (cols 8-10)
        ws.merge_cells("H5:J5")
        ws["H5"] = "Backend"
        ws["H5"].font = Font(bold=True, color="FFFFFF")
        ws["H5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H5"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Frontend (cols 11-14, including Responsive)
        ws.merge_cells("K5:N5")
        ws["K5"] = "Frontend"
        ws["K5"].font = Font(bold=True, color="FFFFFF")
        ws["K5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K5"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Testing (col 15)
        ws["O5"] = "Testing"
        ws["O5"].font = Font(bold=True, color="FFFFFF")
        ws["O5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["O5"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Row 6: Task type sub-headers
        task_types = ["Implement", "FixBug", "Unit Test"]

        # Backend task types (cols 8-10)
        for i, task_type in enumerate(task_types, start=8):
            cell = ws.cell(row=6, column=i)
            cell.value = task_type
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

        # Frontend task types (cols 11-13)
        for i, task_type in enumerate(task_types, start=11):
            cell = ws.cell(row=6, column=i)
            cell.value = task_type
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

        # Frontend Responsive (col 14)
        ws.cell(row=6, column=14).value = "Responsive"
        ws.cell(row=6, column=14).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=6, column=14).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=6, column=14).fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

        # Testing Implement (col 15)
        ws.cell(row=6, column=15).value = "Implement"
        ws.cell(row=6, column=15).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=6, column=15).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=6, column=15).fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

        # Row 7: Main column headers
        for idx, col_name in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=7, column=idx)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF")  # White text for contrast
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Professional blue

            # Auto-calculate column width based on header text length
            # Formula: (character_count * 1.2) + 2 (padding), minimum width = 10
            char_count = len(col_name)
            calculated_width = max(char_count * 1.2 + 2, 10)

            # Apply special minimum widths for specific columns
            if col_name in self.COLUMN_MIN_WIDTHS:
                calculated_width = max(calculated_width, self.COLUMN_MIN_WIDTHS[col_name])

            ws.column_dimensions[get_column_letter(idx)].width = calculated_width

    def _build_data_table(self, ws, data: List[Dict[str, Any]]) -> int:
        """Build data table with task rows, grouped by category.

        Returns:
            int: Last row number used
        """
        start_row = 8  # Data starts at row 8

        # Group tasks by category
        from itertools import groupby

        # Sort by category first to ensure grouping works
        sorted_data = sorted(data, key=lambda x: x.get("category", ""))

        current_row = start_row
        task_number = 1

        # Process each category group
        for category, tasks in groupby(sorted_data, key=lambda x: x.get("category", "")):
            tasks_list = list(tasks)

            # First task in category: show category name
            first_task = True

            for task in tasks_list:
                row = current_row

                # No (auto-increment)
                ws.cell(row=row, column=1).value = task_number
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

                # Category - Only show for first task in group
                if first_task:
                    ws.cell(row=row, column=2).value = category
                    first_task = False
                else:
                    ws.cell(row=row, column=2).value = ""  # Empty for subsequent rows
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")

                # Parent Task
                ws.cell(row=row, column=3).value = task.get("parent_task", "")
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")

                # Sub Task
                ws.cell(row=row, column=4).value = task.get("sub_task", "")
                ws.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="center")

                # Sub.No
                ws.cell(row=row, column=5).value = task.get("sub_no", "")
                ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")

                # Premise
                ws.cell(row=row, column=6).value = task.get("premise", "")
                ws.cell(row=row, column=6).alignment = Alignment(horizontal="left", vertical="center")

                # 備考 Remark
                ws.cell(row=row, column=7).value = task.get("remark", "")
                ws.cell(row=row, column=7).alignment = Alignment(horizontal="left", vertical="center")

                # Effort columns
                backend_data = task.get("backend", {})
                frontend_data = task.get("frontend", {})
                responsive_data = task.get("responsive", {})
                testing_data = task.get("testing", {})

                # Backend - Implement
                ws.cell(row=row, column=8).value = backend_data.get("implement", 0) or None
                ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="center")

                # Backend - FixBug
                ws.cell(row=row, column=9).value = backend_data.get("fixbug", 0) or None
                ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")

                # Backend - Unit Test
                ws.cell(row=row, column=10).value = backend_data.get("unittest", 0) or None
                ws.cell(row=row, column=10).alignment = Alignment(horizontal="center", vertical="center")

                # Frontend - Implement
                ws.cell(row=row, column=11).value = frontend_data.get("implement", 0) or None
                ws.cell(row=row, column=11).alignment = Alignment(horizontal="center", vertical="center")

                # Frontend - FixBug
                ws.cell(row=row, column=12).value = frontend_data.get("fixbug", 0) or None
                ws.cell(row=row, column=12).alignment = Alignment(horizontal="center", vertical="center")

                # Frontend - Unit Test
                ws.cell(row=row, column=13).value = frontend_data.get("unittest", 0) or None
                ws.cell(row=row, column=13).alignment = Alignment(horizontal="center", vertical="center")

                # Frontend - Responsive
                ws.cell(row=row, column=14).value = responsive_data.get("implement", 0) or None
                ws.cell(row=row, column=14).alignment = Alignment(horizontal="center", vertical="center")

                # Testing - Implement
                ws.cell(row=row, column=15).value = testing_data.get("implement", 0) or None
                ws.cell(row=row, column=15).alignment = Alignment(horizontal="center", vertical="center")

                # Total (MD) - Formula: SUM of effort columns
                effort_start = get_column_letter(self.EFFORT_COLUMNS[0])
                effort_end = get_column_letter(self.EFFORT_COLUMNS[-1])
                total_formula = f"=SUM({effort_start}{row}:{effort_end}{row})"
                ws.cell(row=row, column=16).value = total_formula
                ws.cell(row=row, column=16).alignment = Alignment(horizontal="center", vertical="center")

                # Note
                ws.cell(row=row, column=17).value = task.get("note", "")
                ws.cell(row=row, column=17).alignment = Alignment(horizontal="left", vertical="center")

                current_row += 1
                task_number += 1

        # Return last row number (current_row - 1 since we incremented)
        return current_row - 1

    def _add_total_mm_row(self, ws, last_row: int):
        """Add Total (MM) summary row with formulas."""
        total_row = last_row + 1  # Row after last data row

        # Merge cells for "Total (MM)" label
        ws.merge_cells(f"A{total_row}:G{total_row}")
        ws[f"A{total_row}"] = "Total (MM)"
        ws[f"A{total_row}"].font = Font(bold=True)
        ws[f"A{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{total_row}"].fill = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")

        # Add MM formulas for each effort column and Total (MD)
        for col_idx in self.EFFORT_COLUMNS + [self.TOTAL_COLUMN]:
            col_letter = get_column_letter(col_idx)

            # Formula: SUM(column) / $M$1
            mm_formula = f"=SUM({col_letter}8:{col_letter}{total_row-1})/$M$1"

            cell = ws.cell(row=total_row, column=col_idx)
            cell.value = mm_formula
            cell.number_format = "0.00"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")

    def _apply_formatting(self, ws, last_row: int):
        """Apply borders and final formatting."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Apply borders to header rows (rows 5, 6, 7)
        for row_num in [5, 6, 7]:
            for col_idx in range(1, len(self.COLUMNS) + 1):
                ws.cell(row=row_num, column=col_idx).border = thin_border

        # Apply borders to data rows (row 8 to total_row)
        total_row = last_row + 1  # Total MM row
        for row in range(8, total_row + 1):
            for col_idx in range(1, len(self.COLUMNS) + 1):
                ws.cell(row=row, column=col_idx).border = thin_border


def export_sunasterisk_excel(
    data: List[Dict[str, Any]],
    filename: Optional[str] = None,
    no: str = "001",
    version: str = "1.0",
    issue_date: Optional[str] = None,
    md_per_mm: int = 20,
    project_id: str = None
) -> str:
    """
    Convenience function to export Sun Asterisk Excel estimation sheet.

    Args:
        data: List of task dictionaries
        filename: Output filename (auto-generated if None)
        no: Document number
        version: Document version
        issue_date: Issue date (defaults to today)
        md_per_mm: Man-days per man-month
        project_id: Project identifier for project-scoped storage

    Returns:
        str: Path to exported Excel file
    """
    exporter = SunAsteriskExcelExporter(
        no=no,
        version=version,
        issue_date=issue_date,
        md_per_mm=md_per_mm
    )
    return exporter.export(data, filename, project_id=project_id)
